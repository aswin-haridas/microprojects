from PIL import Image
import openpyxl
import os

# Define the colors and their values
colors = {
    'red': (0, 255, 255),
    'orange': (30, 255, 255),
    'yellow': (60, 255, 255),
    'green': (90, 255, 255),
    'aqua': (120, 255, 255),
    'blue': (150, 255, 255),
    'purple': (180, 255, 255),
    'magenta': (300, 255, 255)
}

digital_images_dir = r'T:\\ML\\digital'
film_images_dir = r'T:\\ML\\film'
output_file = 'color_changes.xlsx'

# Create a workbook and select the active sheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write the headers for the columns
worksheet.cell(row=1, column=1, value='Image Name')

column = 2
for color in colors:
    worksheet.cell(row=1, column=column, value=f'{color} - Hue Change')
    worksheet.cell(row=1, column=column + 1, value=f'{color} - Saturation Change')
    worksheet.cell(row=1, column=column + 2, value=f'{color} - Brightness Change')
    column += 3

# Iterate over the digital images directory
for filename in os.listdir(digital_images_dir):
    if filename.endswith('.jpg') or filename.endswith('.png'):
        image_path = os.path.join(digital_images_dir, filename)
        image = Image.open(image_path)
        image_hsv = image.convert('HSV')

        # Write the image name to the worksheet
        row = worksheet.max_row + 1
        worksheet.cell(row=row, column=1, value=filename)

        # Calculate and write the color changes for each color
        column = 2
        for color in colors:
            hsv_color = colors[color]
            pixel_color = image_hsv.getpixel((0, 0))
            hue_change = min(100, max(-100, hsv_color[0] - pixel_color[0]))
            saturation_change = min(100, max(-100, hsv_color[1] - pixel_color[1]))
            brightness_change = min(100, max(-100, hsv_color[2] - pixel_color[2]))

            worksheet.cell(row=row, column=column, value=hue_change)
            worksheet.cell(row=row, column=column + 1, value=saturation_change)
            worksheet.cell(row=row, column=column + 2, value=brightness_change)
            column += 3

# Iterate over the film images directory (similar to the digital images section)

# Save the workbook to the output file
workbook.save(output_file)
